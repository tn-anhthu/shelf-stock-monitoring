"""Split wrong LLM predictions in a ground-truth-filled data/scan_viz/review.xlsx
sheet into two buckets, to tell apart two different failure modes:

  retrieval failure: SigLIP2's cosine-similarity shortlist (top5_candidates)
                      never contained the true SKU at all -- the LLM never
                      had a chance to pick it.
  reasoning failure: the true SKU WAS in top5_candidates, but the LLM still
                      picked a different one -- a prompt/model reasoning
                      problem, not a retrieval problem.

Only rows with correct='no' and a real true_sku_id (not empty, not the
"unknown" sentinel -- see src/pipeline/llm_escalation.py's answer schema)
are considered; those are the only rows where we know both what SigLIP2
shortlisted and what the right answer actually was.

Usage:
    python3 scripts/analyze_recall.py --sheet test1_llm
    python3 scripts/analyze_recall.py --xlsx data/scan_viz/review.xlsx --sheet test1_llm
"""
import argparse
import sys
from pathlib import Path
from typing import Dict, List

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from src.pipeline.review_export import parse_candidates

DEFAULT_XLSX_PATH = "data/scan_viz/review.xlsx"
UNKNOWN_SENTINELS = {"", "unknown"}


def analyze_recall(xlsx_path: str, sheet_name: str) -> Dict[str, List]:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet_name]

    header = [cell.value for cell in ws[1]]
    col = {name: i for i, name in enumerate(header)}

    retrieval_failures = []
    reasoning_failures = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        correct = str(row[col["correct"]] or "").strip().lower()
        true_sku_id = str(row[col["true_sku_id"]] or "").strip()
        if correct != "no" or true_sku_id.lower() in UNKNOWN_SENTINELS:
            continue

        candidate_skus = parse_candidates(row[col["top5_candidates"]] or "")
        index = row[col["index"]]
        if true_sku_id.lower() in {sku.lower() for sku in candidate_skus}:
            reasoning_failures.append(index)
        else:
            retrieval_failures.append(index)

    return {"retrieval_failures": retrieval_failures, "reasoning_failures": reasoning_failures}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", type=str, default=DEFAULT_XLSX_PATH)
    parser.add_argument("--sheet", type=str, required=True)
    args = parser.parse_args()

    result = analyze_recall(args.xlsx, args.sheet)
    retrieval_failures = result["retrieval_failures"]
    reasoning_failures = result["reasoning_failures"]
    total = len(retrieval_failures) + len(reasoning_failures)

    print(f"Sheet '{args.sheet}': {total} wrong-prediction case(s) with known ground truth")
    print(f"  Retrieval failures (true SKU not in top-5): {len(retrieval_failures)}  indices={retrieval_failures}")
    print(f"  Reasoning failures (true SKU in top-5, LLM still wrong): {len(reasoning_failures)}  indices={reasoning_failures}")


if __name__ == "__main__":
    main()
