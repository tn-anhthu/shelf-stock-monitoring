"""Pilot: compare Gemini (gemini-3.5-flash-lite by default -- see
src/pipeline/llm_escalation.py's GEMINI_MODEL_ID for why not 2.5 Flash, and
for the cheaper-but-still-working -lite alternative to 3.6-flash) against
Claude on the SAME crops/candidates, without spending any more Claude API
credit.

Motivation: the project's Claude credit is exhausted, but every crop from a
prior real e2e run is already sitting in data/scan_viz/review.xlsx --
predicted_sku_id (Claude's answer), score, top5_candidates (the exact SigLIP2
shortlist Claude was shown), llm_reasoning, and (once hand-filled)
correct/true_sku_id ground truth. This script reuses all of that instead of
re-running SigLIP2 or calling Claude again: for each row in a given sheet, it
loads the same crop image and the same candidate list Claude saw (via
src.pipeline.review_export.parse_candidates), calls ONLY
escalate_to_llm_gemini fresh, and prints Claude's already-recorded answer
next to Gemini's fresh one -- so the two providers are compared on identical
inputs, spending only Gemini credit.

Usage:
    python3 scripts/pilot_gemini_vs_claude.py --claude-sheet test1_recall2
    python3 scripts/pilot_gemini_vs_claude.py --claude-sheet test1_recall2 --indices 0,6,41

Crop images are read from data/scan_viz/<claude-sheet>/crop_XX_ok.jpg -- the
same output directory scripts/visualize_scan_e2e.py wrote them to (out_dir.name
becomes the xlsx sheet name, so the two always match by convention). If
filter_contained_boxes dropped a different box count between two runs, row
`index` values and crop_XX numbers shift -- always pull crop_file straight
from the sheet being read, never assume a crop number from a different
sheet/run still points at the same product.

API key: GEMINI_API_KEY, read from the environment (or a gitignored .env
file via python-dotenv, same as ANTHROPIC_API_KEY in visualize_scan_e2e.py).
"""
import argparse
import os
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import openpyxl
from google import genai
from PIL import Image

from src.catalog.db import get_connection, list_catalog
from src.pipeline.llm_escalation import escalate_to_llm_gemini
from src.pipeline.review_export import parse_candidates

DEFAULT_XLSX_PATH = "data/scan_viz/review.xlsx"
DEFAULT_DB_PATH = "data/shelfsense.db"
UNKNOWN_SENTINELS = {"", "unknown"}
VERDICT_LABEL = {"yes": "RIGHT", "no": "WRONG", "?": "?"}

# gemini-3.5-flash-lite pricing (https://ai.google.dev/gemini-api/docs/pricing,
# checked 2026-07-27), per million tokens -- used to turn this run's summed
# usage into a rough dollar estimate, same approach as visualize_scan_e2e.py's
# Haiku cost constants. Update these if GEMINI_ESCALATION_MODEL changes --
# e.g. gemini-3.6-flash is $1.50/$7.50, a full (non-lite) tier confirmed
# working if quality matters more than cost.
GEMINI_INPUT_COST_PER_MTOK = 0.30
GEMINI_OUTPUT_COST_PER_MTOK = 2.50


def _parse_indices(raw: str) -> List[int]:
    return [int(part.strip()) for part in raw.split(",") if part.strip()]


def _gemini_verdict(predicted_sku_id: Optional[str], true_sku_id: str) -> str:
    if not true_sku_id or true_sku_id.lower() in UNKNOWN_SENTINELS:
        return "?"
    predicted = (predicted_sku_id or "").strip().lower()
    return "RIGHT" if predicted == true_sku_id.strip().lower() else "WRONG"


def load_sheet_rows(xlsx_path: str, sheet_name: str, indices: Optional[List[int]] = None) -> List[Dict]:
    """Reads a review.xlsx sheet and returns one dict per crop worth testing --
    skipping rows with no saved crop image (crop_status was "degenerate") or
    no top5_candidates (SigLIP2 never produced a shortlist, so there's
    nothing for Gemini to choose among either). `correct` is normalized to
    lowercase "yes"/"no"/"?" since it's hand-typed into Excel."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet_name]
    header = [cell.value for cell in ws[1]]
    col = {name: i for i, name in enumerate(header)}

    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        index = raw[col["index"]]
        if indices is not None and index not in indices:
            continue
        crop_file = raw[col["crop_file"]]
        top5_candidates = raw[col["top5_candidates"]] or ""
        if not crop_file or not top5_candidates:
            continue
        rows.append({
            "index": index,
            "crop_file": crop_file,
            "predicted_sku_id": raw[col["predicted_sku_id"]],
            "true_sku_id": str(raw[col["true_sku_id"]] or "").strip(),
            "claude_correct": str(raw[col["correct"]] or "").strip().lower() or "?",
            "candidate_sku_ids": parse_candidates(top5_candidates),
        })
    return rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--claude-sheet", type=str, required=True)
    parser.add_argument("--xlsx", type=str, default=DEFAULT_XLSX_PATH)
    parser.add_argument("--db", type=str, default=DEFAULT_DB_PATH)
    parser.add_argument(
        "--indices", type=str, default=None,
        help="Comma-separated row `index` values to limit the run to (saves Gemini credit); default: every row in the sheet",
    )
    parser.add_argument(
        "--delay-seconds", type=float, default=13.0,
        help=(
            "Seconds to sleep between Gemini calls -- the free tier caps gemini-3.6-flash at "
            "5 requests/minute (verified live: a real run hit RESOURCE_EXHAUSTED after 6 calls), "
            "so the default paces to roughly 60/5=12s + margin as a conservative default across "
            "models. Set lower (e.g. 0) on a paid tier / once billing is enabled."
        ),
    )
    args = parser.parse_args()

    if not os.environ.get("GEMINI_API_KEY"):
        raise SystemExit("GEMINI_API_KEY not set — export it before running this script.")
    gemini_client = genai.Client(api_key=os.environ["GEMINI_API_KEY"])

    conn = get_connection(args.db)
    catalog_items = list_catalog(conn)
    names_by_sku = {item["sku_id"]: item.get("name", item["sku_id"]) for item in catalog_items}

    indices = _parse_indices(args.indices) if args.indices else None
    rows = load_sheet_rows(args.xlsx, args.claude_sheet, indices=indices)
    crops_dir = Path("data/scan_viz") / args.claude_sheet

    total_input_tokens, total_output_tokens = 0, 0
    claude_right, gemini_right, compared = 0, 0, 0

    calls_made = 0
    for row in rows:
        crop_path = crops_dir / row["crop_file"]
        if not crop_path.exists():
            print(f"  skip index {row['index']}: {crop_path} not found")
            continue
        image = Image.open(crop_path)
        candidates = [(sku_id, names_by_sku.get(sku_id, sku_id)) for sku_id in row["candidate_sku_ids"]]

        if calls_made > 0 and args.delay_seconds > 0:
            time.sleep(args.delay_seconds)
        gemini_answer, gemini_reasoning, usage = escalate_to_llm_gemini(gemini_client, image, candidates)
        calls_made += 1
        total_input_tokens += usage["input_tokens"]
        total_output_tokens += usage["output_tokens"]
        cost = (
            usage["input_tokens"] / 1_000_000 * GEMINI_INPUT_COST_PER_MTOK
            + usage["output_tokens"] / 1_000_000 * GEMINI_OUTPUT_COST_PER_MTOK
        )

        claude_verdict = VERDICT_LABEL.get(row["claude_correct"], "?")
        gemini_verdict = _gemini_verdict(gemini_answer, row["true_sku_id"])

        compared += 1
        if row["claude_correct"] == "yes":
            claude_right += 1
        if gemini_verdict == "RIGHT":
            gemini_right += 1

        print(
            f"[{row['index']:02d}] {row['crop_file']:<20s} true={row['true_sku_id'] or '?':<25s} "
            f"claude={str(row['predicted_sku_id'] or 'unknown'):<25s}({claude_verdict})  "
            f"gemini={gemini_answer or 'unknown':<25s}({gemini_verdict})  "
            f"tokens={usage['input_tokens']}/{usage['output_tokens']} (${cost:.4f})"
        )
        print(f"       gemini_reasoning: {gemini_reasoning}")

    estimated_cost = (
        total_input_tokens / 1_000_000 * GEMINI_INPUT_COST_PER_MTOK
        + total_output_tokens / 1_000_000 * GEMINI_OUTPUT_COST_PER_MTOK
    )
    print(f"\nCompared {compared} crop(s) from sheet '{args.claude_sheet}'")
    print(f"Claude accuracy (from sheet's hand-filled 'correct' column): {claude_right}/{compared}")
    print(f"Gemini accuracy (fresh, vs. sheet's true_sku_id): {gemini_right}/{compared}")
    print(
        f"Gemini token usage: {total_input_tokens} input, {total_output_tokens} output "
        f"-> estimated cost ${estimated_cost:.4f}"
    )


if __name__ == "__main__":
    main()
