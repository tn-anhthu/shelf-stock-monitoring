"""Re-run only the LLM verification step (src/pipeline/classify.py::verify_with_llm)
for specific already-scanned rows in data/scan_viz/review.xlsx, reusing each row's
existing top5_candidates shortlist -- no re-detect/re-embed/re-rank against SigLIP2.

For rows whose predicted_sku_id is empty because the LLM call itself errored
(e.g. Gemini 503 UNAVAILABLE) rather than genuinely answering "unknown",
updates predicted_sku_id, score, and llm_reasoning in place for exactly the
given rows -- every other row, and the correct/true_sku_id ground-truth
columns, are left untouched.

Reads each row's crop image from <scan-viz-dir>/<sheet>/<crop_file> (the same
crop files scripts/visualize_scan_e2e.py already saved for that run).

Provider/API key/fallback picked the same way as the other e2e scripts (see
src/pipeline/classify.py::_escalate) -- LLM_PROVIDER env var, GEMINI_API_KEY
or ANTHROPIC_API_KEY, and (gemini only) an automatic OpenRouter retry via
CLASSIFY_FALLBACK_MODEL if the primary call errors.

Usage:
    python3 scripts/reverify_review_rows.py --targets test20:0,test20:8,test21:2
"""
import argparse
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import anthropic
import openpyxl
from google import genai
from PIL import Image
import pillow_heif
pillow_heif.register_heif_opener()

from src.catalog.db import get_connection, list_catalog
from src.pipeline.classify import verify_with_llm
from src.pipeline.review_export import COLUMNS

REVIEW_XLSX_PATH = "data/scan_viz/review.xlsx"
DEFAULT_DB_PATH = "data/shelfsense.db"
DEFAULT_SCAN_VIZ_DIR = "data/scan_viz"


def _parse_targets(text: str):
    targets = []
    for chunk in text.split(","):
        sheet, index = chunk.strip().split(":")
        targets.append((sheet.strip(), int(index)))
    return targets


def _parse_candidates_with_scores(text: str):
    """Inverse of review_export.format_candidates, keeping the score (unlike
    review_export.parse_candidates, which drops it) -- verify_with_llm needs
    each candidate's own cosine score to report back as `score` when the LLM
    picks something other than the top1 candidate."""
    if not text:
        return []
    pairs = []
    for chunk in text.split(","):
        sku_id, score = chunk.strip().split(":")
        pairs.append((sku_id.strip(), float(score)))
    return pairs


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--targets", required=True, help="comma-separated sheet:index pairs, e.g. test20:0,test20:8")
    parser.add_argument("--db", default=DEFAULT_DB_PATH)
    parser.add_argument("--scan-viz-dir", default=DEFAULT_SCAN_VIZ_DIR)
    args = parser.parse_args()

    targets = _parse_targets(args.targets)

    provider = os.environ.get("LLM_PROVIDER", "anthropic")
    if provider == "gemini":
        if not os.environ.get("GEMINI_API_KEY"):
            raise SystemExit("GEMINI_API_KEY not set — export it before running this script.")
        llm_client = genai.Client(api_key=os.environ["GEMINI_API_KEY"])
    else:
        if not os.environ.get("ANTHROPIC_API_KEY"):
            raise SystemExit("ANTHROPIC_API_KEY not set — export it before running this script.")
        llm_client = anthropic.Anthropic()

    conn = get_connection(args.db)
    catalog_items = list_catalog(conn)

    wb = openpyxl.load_workbook(REVIEW_XLSX_PATH)
    col = {name: i for i, name in enumerate(COLUMNS)}

    for sheet_name, index in targets:
        if sheet_name not in wb.sheetnames:
            print(f"{sheet_name}[{index}]: sheet not found, skipping")
            continue
        ws = wb[sheet_name]
        target_row = next((row for row in ws.iter_rows(min_row=2) if row[col["index"]].value == index), None)
        if target_row is None:
            print(f"{sheet_name}[{index}]: row not found, skipping")
            continue

        crop_file = target_row[col["crop_file"]].value
        ranked = _parse_candidates_with_scores(target_row[col["top5_candidates"]].value)
        crop_path = Path(args.scan_viz_dir) / sheet_name / crop_file
        crop_image = Image.open(crop_path) if crop_file else None

        sku_id, score, reasoning, _usage, _ranked = verify_with_llm(crop_image, ranked, catalog_items, llm_client)

        target_row[col["predicted_sku_id"]].value = sku_id
        target_row[col["score"]].value = score
        target_row[col["llm_reasoning"]].value = reasoning

        tag = sku_id or "UNKNOWN"
        print(f"{sheet_name}[{index}] ({crop_file}): {tag}  score={score:.3f}  reasoning={reasoning[:100]}")

    wb.save(REVIEW_XLSX_PATH)
    print(f"\nSaved {len(targets)} updated row(s) to {REVIEW_XLSX_PATH}")


if __name__ == "__main__":
    main()
