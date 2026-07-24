import openpyxl
import pytest

from src.pipeline.review_export import append_review_sheet, format_candidates, parse_candidates


def make_rows():
    return [
        {
            "index": 0,
            "crop_file": "crop_00_ok.jpg",
            "predicted_sku_id": "choco_pie_orion",
            "score": 0.91,
            "top5_candidates": "choco_pie_orion:0.91, coke_330:0.55",
            "llm_reasoning": "Logo và bao bì khớp hoàn toàn",
            "correct": "",
            "true_sku_id": "",
            "depth": 1,
        },
        {
            "index": 1,
            "crop_file": "crop_01_ok.jpg",
            "predicted_sku_id": None,
            "score": 0.42,
            "top5_candidates": "coke_330:0.42, choco_pie_orion:0.38",
            "llm_reasoning": "Không thấy candidate nào khớp thương hiệu",
            "correct": "",
            "true_sku_id": "",
            "depth": 1,
        },
    ]


def test_append_review_sheet_creates_new_file_when_missing(tmp_path):
    xlsx_path = tmp_path / "review.xlsx"

    append_review_sheet(str(xlsx_path), "test1_llm", make_rows())

    assert xlsx_path.exists()
    wb = openpyxl.load_workbook(xlsx_path)
    assert wb.sheetnames == ["test1_llm"]
    ws = wb["test1_llm"]
    assert [c.value for c in ws[1]] == [
        "index", "crop_file", "predicted_sku_id", "score", "top5_candidates", "llm_reasoning", "correct", "true_sku_id", "depth",
    ]
    assert [c.value for c in ws[2]] == [
        0, "crop_00_ok.jpg", "choco_pie_orion", 0.91, "choco_pie_orion:0.91, coke_330:0.55",
        "Logo và bao bì khớp hoàn toàn", None, None, 1,
    ]
    assert [c.value for c in ws[3]] == [
        1, "crop_01_ok.jpg", None, 0.42, "coke_330:0.42, choco_pie_orion:0.38",
        "Không thấy candidate nào khớp thương hiệu", None, None, 1,
    ]


def test_append_review_sheet_adds_sheet_without_touching_existing_ones(tmp_path):
    xlsx_path = tmp_path / "review.xlsx"

    append_review_sheet(str(xlsx_path), "test1_llm", make_rows())

    # Simulate hand-filled ground truth in the first sheet (correct=F, true_sku_id=G).
    wb = openpyxl.load_workbook(xlsx_path)
    wb["test1_llm"]["F2"] = "yes"
    wb["test1_llm"]["G2"] = "choco_pie_orion"
    wb.save(xlsx_path)

    append_review_sheet(str(xlsx_path), "test2_llm", make_rows())

    wb = openpyxl.load_workbook(xlsx_path)
    assert wb.sheetnames == ["test1_llm", "test2_llm"]
    assert wb["test1_llm"]["F2"].value == "yes"
    assert wb["test1_llm"]["G2"].value == "choco_pie_orion"


def test_append_review_sheet_avoids_overwriting_same_sheet_name(tmp_path):
    xlsx_path = tmp_path / "review.xlsx"

    append_review_sheet(str(xlsx_path), "test1_llm", make_rows())
    wb = openpyxl.load_workbook(xlsx_path)
    wb["test1_llm"]["F2"] = "yes"
    wb.save(xlsx_path)

    used_name = append_review_sheet(str(xlsx_path), "test1_llm", make_rows())

    assert used_name != "test1_llm"
    wb = openpyxl.load_workbook(xlsx_path)
    assert wb["test1_llm"]["F2"].value == "yes"
    assert used_name in wb.sheetnames


def test_format_candidates_formats_sku_score_pairs():
    ranked = [("choco_pie_orion", 0.923), ("coke_330", 0.681)]
    assert format_candidates(ranked) == "choco_pie_orion:0.92, coke_330:0.68"


def test_format_candidates_empty_list_returns_empty_string():
    assert format_candidates([]) == ""


def test_parse_candidates_recovers_sku_id_list():
    assert parse_candidates("choco_pie_orion:0.92, coke_330:0.68") == ["choco_pie_orion", "coke_330"]


def test_parse_candidates_empty_string_returns_empty_list():
    assert parse_candidates("") == []


def test_append_review_sheet_raises_clear_error_when_file_locked(tmp_path, monkeypatch):
    xlsx_path = tmp_path / "review.xlsx"
    append_review_sheet(str(xlsx_path), "test1_llm", make_rows())

    def locked_save(self, filename):
        raise PermissionError("locked")

    monkeypatch.setattr(openpyxl.Workbook, "save", locked_save)

    with pytest.raises(RuntimeError, match="đang mở"):
        append_review_sheet(str(xlsx_path), "test2_llm", make_rows())


def test_append_review_sheet_does_not_disturb_a_pre_existing_old_format_sheet(tmp_path):
    # Simulates a real sheet written before llm_reasoning existed (7 columns,
    # e.g. data/scan_viz/review.xlsx's test1_llm/test1_llm_2/test1_llm_3
    # sheets from before this change) — appending a NEW sheet with the
    # current (8-column) COLUMNS must not touch or corrupt it.
    xlsx_path = tmp_path / "review.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "old_format_sheet"
    ws.append(["index", "crop_file", "predicted_sku_id", "score", "correct", "true_sku_id", "depth"])
    ws.append([0, "crop_00_ok.jpg", "choco_pie_orion", 0.91, "yes", "choco_pie_orion", 1])
    wb.save(xlsx_path)

    append_review_sheet(str(xlsx_path), "new_format_sheet", make_rows())

    wb = openpyxl.load_workbook(xlsx_path)
    assert set(wb.sheetnames) == {"old_format_sheet", "new_format_sheet"}

    old_ws = wb["old_format_sheet"]
    assert [c.value for c in old_ws[1]] == [
        "index", "crop_file", "predicted_sku_id", "score", "correct", "true_sku_id", "depth",
    ]
    assert [c.value for c in old_ws[2]] == [0, "crop_00_ok.jpg", "choco_pie_orion", 0.91, "yes", "choco_pie_orion", 1]

    new_ws = wb["new_format_sheet"]
    assert [c.value for c in new_ws[1]] == [
        "index", "crop_file", "predicted_sku_id", "score", "top5_candidates", "llm_reasoning", "correct", "true_sku_id", "depth",
    ]
