"""Append one review sheet per real (LLM-verified) scan run to a single
review workbook (data/scan_viz/review.xlsx), instead of a separate CSV per
run — so ground truth can be hand-filled per sheet and accumulate across
runs in one file. Never overwrites an existing sheet: it may already have
hand-filled ground truth (correct/true_sku_id columns).
"""
import os
from typing import Dict, List, Tuple

import openpyxl

COLUMNS = [
    "index", "crop_file", "predicted_sku_id", "score", "top5_candidates", "llm_reasoning", "correct", "true_sku_id",
    "depth",
]

LOCKED_FILE_MESSAGE = "{path} đang mở, đóng file rồi chạy lại."


def format_candidates(ranked: List[Tuple[str, float]]) -> str:
    """Format a rank_candidates()-style (sku_id, score) shortlist into the
    top5_candidates cell text, e.g. "sku_a:0.72, sku_b:0.68" — parse_candidates
    reverses this for scripts/analyze_recall.py's retrieval-recall check."""
    return ", ".join(f"{sku_id}:{score:.2f}" for sku_id, score in ranked)


def parse_candidates(text: str) -> List[str]:
    """Parse a top5_candidates cell back into just its sku_id list — the
    inverse of format_candidates, used by scripts/analyze_recall.py to check
    whether a ground-truth true_sku_id was in the shortlist SigLIP2 handed
    the LLM."""
    if not text:
        return []
    return [pair.split(":", 1)[0].strip() for pair in text.split(",")]


def _unique_sheet_name(existing_names: List[str], sheet_name: str) -> str:
    if sheet_name not in existing_names:
        return sheet_name
    n = 2
    while f"{sheet_name}_{n}" in existing_names:
        n += 1
    return f"{sheet_name}_{n}"


def append_review_sheet(xlsx_path: str, sheet_name: str, rows: List[Dict]) -> str:
    """Create xlsx_path if missing, else load and add a new sheet to it.

    Returns the sheet name actually used — suffixed with _2, _3, ... if
    sheet_name already exists in the workbook, so a re-run never clobbers a
    sheet that may already have hand-filled ground truth.
    """
    try:
        if os.path.exists(xlsx_path):
            wb = openpyxl.load_workbook(xlsx_path)
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
    except PermissionError:
        raise RuntimeError(LOCKED_FILE_MESSAGE.format(path=xlsx_path))

    actual_sheet_name = _unique_sheet_name(wb.sheetnames, sheet_name)
    ws = wb.create_sheet(title=actual_sheet_name)
    ws.append(COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in COLUMNS])

    try:
        wb.save(xlsx_path)
    except PermissionError:
        raise RuntimeError(LOCKED_FILE_MESSAGE.format(path=xlsx_path))

    return actual_sheet_name
